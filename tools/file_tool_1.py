# coding: utf-8
from config import base_path, cell_config, open_api
import os
import openpyxl
from tools.mylog import Logger
import datetime
from scripts.init_env import current_env
import csv
import pandas as pd

import xlrd
from xlutils.copy import copy
import config

log = Logger()  # 初始化日志对象


class FileTool:
    # 初始化
    def __init__(self, filename, device_type, is_back=True):
        # 组装动态文件路径
        self.old_filename = base_path + os.sep + "data" + os.sep + filename  # 用例文件目录
        if os.path.isfile(self.old_filename) == False:
            raise Exception(f"文件{self.old_filename}不存在")
        nowtimeinfo = datetime.datetime.now().strftime('%Y-%m-%d_%H-%M-%S')
        self.filename = base_path + os.sep + "result" + os.sep + current_env + "_" + device_type + "_" + nowtimeinfo + filename.replace(
            ".csv", ".xlsx")  # 保存的文件名称
        # self.back_excel()  # 备份用例

        self.csv_to_xlsx_pd()
        self.load_excel()  # 加载用例

    def csv_to_xlsx_pd(self):
        csv = pd.read_csv(self.old_filename, encoding='gbk')
        csv.to_excel(self.filename, sheet_name='data', index=0)

    # 备份excel文件
    def load_excel(self):
        # 打开文件，获取workbook对象
        # oldworkbook = openpyxl.load_workbook(self.old_filename)
        # oldworkbook.save(self.filename)
        # oldworkbook.close()
        self.workbook = openpyxl.load_workbook(self.filename)

        # 3、获取所有的表单对象
        self.sheets = self.workbook.sheetnames
        # 4、获取当前表的表单对象
        self.sheet = self.workbook[self.sheets[0]]

    def read_csv(self):
        wholedictinfo = list()
        try:
            log.info("读取用例文件........")
            allsteps = []
            dictinfo = []
            with open(r"D:\project\auto_test\data\data_case.csv", encoding="utf-8") as f:
                reader = csv.reader(f)
                readerlist = list(reader)
            row = 0
            for i in readerlist:
                row = row + 1
                if row == 1:
                    continue
                caseid = i[cell_config.get("case_id")]  # 用例编号信息
                casetitle = i[cell_config.get("case_name")]  # 用例名称信息
                if caseid.strip() != "":
                    allsteps = []
                    dictinfo = {"case_id": caseid, "case_name": casetitle, "case_category": "", "steps": []}
                    wholedictinfo.append(dictinfo)
                linesinfo = dict()
                params_value = i[cell_config.get("params")]  # 参数信息
                if params_value.strip() == "":
                    continue
                linesinfo["step"] = i[cell_config.get("step")]
                linesinfo["params"] = params_value
                linesinfo["x_y"] = [row, cell_config.get("result")]
                linesinfo["x_y_desc"] = [row, cell_config.get("desc")]
                allsteps.append(linesinfo)
                if "steps" in dictinfo:
                    dictinfo["steps"] = allsteps
            return wholedictinfo
        except Exception as e:
            print("异常信息如下：", e)
            return wholedictinfo

    def write_csv(self, row):
        with open(self.filename, 'w', newline='') as f:
            writer = csv.writer(f)
            writer.writerow(row)

    def read_excel_openapi(self):
        wholedictinfo = list()
        try:
            log.info("读取用例文件........")
            allsteps = []
            dictinfo = []
            worksheet = self.sheet
            # print( worksheet.max_row)
            for i in range(2, worksheet.max_row + 1):
                caseid = worksheet.cell(row=i, column=open_api.get("case_id")).value  # 用例编号信息
                Interface_name = worksheet.cell(row=i, column=open_api.get("Interface_name")).value  # 接口信息
                case_name = worksheet.cell(row=i, column=open_api.get("case_name")).value  # 用例名称信息
                serviceUrl = worksheet.cell(row=i, column=open_api.get("serviceUrl")).value  # 用例名称信息
                data = worksheet.cell(row=i, column=open_api.get("data")).value  # 用例名称信息
                expect = worksheet.cell(row=i, column=open_api.get("expect")).value  # 用例名称信息
                data = {"serviceUrl": serviceUrl, "data": eval(data)}
                dictinfo = {"caseid": caseid, "Interface_name": Interface_name, "case_name": case_name, "data": data,
                            "expect": eval(expect), "response": [i, open_api.get("response")],
                            "result": [i, open_api.get("result")]}
                wholedictinfo.append(dictinfo)
            return wholedictinfo
        except Exception as e:
            log.info("读取用例文件异常,异常信息为:{}".format(e))
            raise e

            # return wholedictinfo

    # 读取excel文件
    def read_excel(self):
        wholedictinfo = list()
        # try:
        log.info("读取用例文件........")
        allsteps = []
        dictinfo = []
        worksheet = self.sheet
        for i in range(2, worksheet.max_row + 1):
            case_category = worksheet.cell(row=i, column=cell_config.get("case_category")).value  # 用例场景
            caseid = worksheet.cell(row=i, column=cell_config.get("case_id")).value  # 用例编号信息
            casetitle = worksheet.cell(row=i, column=cell_config.get("case_name")).value  # 用例名称信息
            lock_device = worksheet.cell(row=i, column=cell_config.get("lock_device")).value  # 设备锁
            is_wait = worksheet.cell(row=i, column=cell_config.get("is_wait")).value
            if caseid:
                allsteps = []
                dictinfo = {"case_id": caseid, "case_name": casetitle, "case_category": case_category,
                            "lock_device": lock_device, "is_wait": is_wait, "steps": []}
                if case_category in config.test_category:  # 筛选用例
                    wholedictinfo.append(dictinfo)
            linesinfo = dict()
            params_value = worksheet.cell(row=i, column=cell_config.get("params")).value
            if params_value == None:
                continue
            linesinfo["step"] = worksheet.cell(row=i, column=cell_config.get("step")).value
            linesinfo["params"] = params_value

            linesinfo["x_y"] = [i, cell_config.get("result")]
            linesinfo["x_y_desc"] = [i, cell_config.get("desc")]
            allsteps.append(linesinfo)
            if "steps" in dictinfo:
                dictinfo["steps"] = allsteps
        log.info("读取用例文件完成........")
        return wholedictinfo

    # 写入exel文件
    # x_y 是一个列表
    def write_excel(self, sheet_name, x_y, msg):
        # x_y 参数的格式为列表，如[2,5]
        # self.sheet = self.workbook[sheet_name]
        try:
            self.sheet.cell(x_y[0], x_y[1]).value = msg
        except Exception as e:
            self.sheet.cell(x_y[0], x_y[1]).value = e
        finally:
            # 保存excel
            self.workbook.save(self.filename)


class MyXlrs:
    def __init__(self, filepath):
        self.filepath = filepath
        self.workbook = xlrd.open_workbook(filepath)

    def get_sheet_names(self):
        sheet_names = self.workbook.sheet_names()
        return sheet_names

    def read_cellvalue(self, row, col, bookname=None):  # row  行  col 列
        if bookname == None:
            bookname = self.get_sheet_names()[0]
        shell_obj = self.workbook.sheet_by_name(bookname)
        result = shell_obj.cell_value(row, col)
        return result

    def read_xlr(self, sheetname=None, start_line=None):
        if start_line == None:
            start_line = 1
        if sheetname == None:
            sheetname = self.get_sheet_names()[0]
        wholedictinfo = list()
        # try:
        log.info("读取用例文件........")
        allsteps = []
        dictinfo = []
        worksheet = self.workbook.sheet_by_name(sheetname)
        # print( worksheet.nrows)
        for i in range(start_line, worksheet.nrows):
            case_category = worksheet.cell_value(rowx=i, colx=cell_config.get("case_category") - 1)  # 用例场景
            col = cell_config.get("case_id")
            caseid = worksheet.cell_value(rowx=i, colx=col - 1)  # 用例编号信息
            casetitle = worksheet.cell_value(rowx=i, colx=cell_config.get("case_name") - 1)  # 用例名称信息
            lock_device = worksheet.cell_value(rowx=i, colx=cell_config.get("lock_device") - 1)  # 用例名称信息

            if caseid:
                allsteps = []
                dictinfo = {"case_id": caseid, "case_name": casetitle, "case_category": case_category,
                            "lock_device": lock_device, "steps": []}
                if case_category in config.test_category:  # 筛选用例
                    wholedictinfo.append(dictinfo)

            linesinfo = dict()
            params_value = worksheet.cell_value(rowx=i, colx=cell_config.get("params") - 1)
            # print(params_value)
            if params_value == None:
                continue
            linesinfo["step"] = worksheet.cell_value(rowx=i, colx=cell_config.get("step") - 1)
            linesinfo["params"] = params_value
            linesinfo["x_y"] = [i, cell_config.get("result")]
            linesinfo["x_y_desc"] = [i, cell_config.get("desc")]
            linesinfo["step_result"] = worksheet.cell_value(rowx=i, colx=cell_config.get("step_result") - 1)
            allsteps.append(linesinfo)
            if "steps" in dictinfo:
                dictinfo["steps"] = allsteps
        log.info("读取用例文件完成........")
        return wholedictinfo

    def copy_sheet(self):
        sheet = copy(self.workbook)
        return sheet

    def write_onedata(self, sheet, x_y, value, result_path=None, sheetname=None):
        if sheetname == None:
            index = 0
        else:
            index = self.get_sheet_names().index(sheetname)
        if result_path == None:
            result_path = self.filepath
        load_sheet = sheet.get_sheet(index)
        load_sheet.write(x_y[0], x_y[1] - 1, value)
        sheet.save(result_path)

    # def save_write(self, w, new_path):
    #     try:
    #         w.save(new_path)
    #     except:
    #         os.makedirs(os.path.dirname(new_path))
    #         w.save(new_path)


if __name__ == '__main__':
    # # result_path=r"F:\git\Midea\auto_test\result\sit_328_2021-01-28data_case.xlsx"
    # # r = MyXlrs(a)
    # # d = r.read_xlr()
    # r = FileTool("open_api_case.csv", "OPEN_API")
    # d = r.read_excel_openapi()
    # print(d)
    # # w_sheet=r.copy_sheet()
    # # r.write_onedata(w_sheet, [1, 8], "test_data23",)
    # # r.save_write(w_sheet, a)
    device_type = "328_halfDuplex"
    a = r"F:\git\Midea\auto_test\data\空调.csv"
    tool = FileTool("空调.csv", device_type)
    # tool.load_excel()
    # 读取excel的内容信息
    testcaseinfo = tool.read_excel()
    print(testcaseinfo)
